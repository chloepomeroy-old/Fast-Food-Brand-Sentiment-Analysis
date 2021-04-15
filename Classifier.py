#!/usr/bin/env python
# coding: utf-8

from Sentiment_Analysis_Model import *

from nltk.tokenize import word_tokenize
from openpyxl import Workbook, load_workbook
import emoji
import pandas as pd

data = load_workbook('twitter_replies.xlsx')
columns = next(data['Sheet'].values)[0:]
df = pd.DataFrame(data['Sheet'].values, columns=columns)
df = df.drop([df.index[0]]) #drop the original header row

column_names = ['date', 'time', 'brand', 'brand_post_url', 'user', 'user name', 'reply', 'sentiment']
wb=Workbook()
page=wb.active
page.append(column_names) #append the first line in the file with the column names

#For each tweet in our scraped tweets, run the sentiment classifier
for index, row in df.iterrows():
    user_reply = row['text']
    brand = row['brand']
    brand_post_url = row['brand_post_url']
    user = row['user']
    user_name = row['user name']
    date = row['date']
    time = row['time']
    reply_tokens = remove_noise(word_tokenize(emoji.demojize((user_reply)))) #remove noise and tokenize tweets
    sentiment = [date, time, brand, brand_post_url, user, user_name, user_reply, classifier.classify(dict([token, True] for token in reply_tokens))]
    page.append(sentiment) #append a line in the file with the tweet and sentiment result    

wb.save(filename = 'twitter_sentiments.xlsx')
